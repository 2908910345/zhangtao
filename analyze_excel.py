import openpyxl
import pandas as pd

filepath = r"d:\代码\刷底稿\引出列表_凭证_0512141615.xlsx"

# ========== 1. 用 openpyxl 读取 sheet 名 ==========
wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
print("=" * 80)
print("📋 文件信息")
print("=" * 80)
print(f"文件名: 引出列表_凭证_0512141615.xlsx")
print(f"Sheet 数量: {len(wb.sheetnames)}")
print(f"Sheet 名称列表: {wb.sheetnames}")
print()

# ========== 2. 遍历每个 sheet 输出结构 ==========
for sname in wb.sheetnames:
    ws = wb[sname]
    print("=" * 80)
    print(f"📄 Sheet: [{sname}]")
    print(f"   行数 (approx): {ws.max_row}, 列数 (approx): {ws.max_column}")
    print("-" * 80)

    # 读取前几行
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append(row)
        if i >= 9:  # 读 10 行
            break

    if not rows:
        print("   (空 sheet)")
        print()
        continue

    # 取第1行作为列名候选
    header = rows[0]
    print(f"  ▶ 第1行 (候选列名, 共 {len(header)} 列):")
    for idx, val in enumerate(header, start=1):
        col_letter = openpyxl.utils.get_column_letter(idx)
        print(f"     列 {col_letter} ({idx}): {repr(val)}")
    print()

    # 输出后续行数据
    print(f"  ▶ 第2~{len(rows)}行 数据预览:")
    for ri, row in enumerate(rows[1:], start=2):
        vals = []
        for idx, val in enumerate(row):
            col_letter = openpyxl.utils.get_column_letter(idx + 1)
            if val is None:
                continue
            vals.append(f"{col_letter}={repr(val)}")
        print(f"     行{ri}: {', '.join(vals)}")
    print()

wb.close()

# ========== 3. 用 pandas 再读一遍，确认列名和数据 ==========
print("=" * 80)
print("🐼 pandas 读取确认")
print("=" * 80)
xls = pd.ExcelFile(filepath, engine="openpyxl")
for sname in xls.sheet_names:
    print(f"\n--- Sheet: {sname} ---")
    df = pd.read_excel(xls, sheet_name=sname, nrows=10, engine="openpyxl")
    print(f"   shape: {df.shape}")
    print(f"   列名: {list(df.columns)}")
    print(f"   各列数据类型:")
    for col in df.columns:
        non_null_cnt = df[col].notna().sum()
        sample_vals = df[col].dropna().unique()[:5].tolist()
        print(f"     {col}: dtype={df[col].dtype}, non_null={non_null_cnt}, samples={sample_vals}")
    print(f"\n   前3行数据:")
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', 200)
    pd.set_option('display.max_colwidth', 40)
    print(df.head(3).to_string(index=True))

xls.close()
print("\n✅ 分析完成")
