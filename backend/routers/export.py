from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl
from urllib.parse import quote
from backend.database import async_session_factory, escape_like
from backend.models import BalanceSubject
from sqlalchemy import select, func

router = APIRouter(prefix="/api", tags=["模板下载与底稿导出"])


def _download_excel(stream, filename):
    encoded = quote(filename)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}",
        }
    )


@router.get("/templates/balance")
async def download_balance_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "科目余额表"
    ws.append(["科目编码", "科目名称", "年初余额-借方", "年初余额-贷方",
                "本期发生额-借方", "本期发生额-贷方",
                "本年累计-借方", "本年累计-贷方",
                "期末余额-借方", "期末余额-贷方", "核算维度"])
    ws.append(["1002", "银行存款", 100000, "", 50000, 20000, 50000, 20000, 130000, "", "银行账户:XX支行"])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return _download_excel(stream, "科目余额表模板.xlsx")


@router.get("/templates/journal")
async def download_journal_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "序时账"
    ws.append(["核算组织", "期间", "凭证号", "摘要", "科目编码", "科目名称", "借方", "贷方", "核算维度"])
    ws.append(["XX公司", "2026年1期", "0001", "计提工资", "6602.18", "管理费用_审计咨询费", 10000, "", "部门:综合部"])
    ws.append(["XX公司", "2026年1期", "0001", "计提工资", "2211.02", "应付职工薪酬_奖金", "", 10000, ""])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return _download_excel(stream, "序时账模板.xlsx")


@router.get("/export/draft")
async def export_balance_to_draft(
    book_name: str = Query("default"),
    code_prefix: str = Query(""),
    level: int = Query(0),
):
    async with async_session_factory() as session:
        query = select(
            BalanceSubject.code, BalanceSubject.name,
            BalanceSubject.year_start_debit, BalanceSubject.year_start_credit,
            BalanceSubject.period_debit, BalanceSubject.period_credit,
            BalanceSubject.year_total_debit, BalanceSubject.year_total_credit,
            BalanceSubject.end_debit, BalanceSubject.end_credit,
            BalanceSubject.dimension,
        ).where(BalanceSubject.book_name == book_name)

        if code_prefix:
            query = query.where(BalanceSubject.code.like(f"{escape_like(code_prefix)}%", escape="/"))

        if level > 0:
            # level=1 表示1级科目（无点号），level=2 表示2级科目（1个点号），以此类推
            target_dots = level - 1
            if target_dots == 0:
                query = query.where(
                    ~BalanceSubject.code.like("%.%"),
                )
            else:
                dot_pattern = "." * target_dots
                query = query.where(
                    BalanceSubject.code.like(f"%{dot_pattern}%"),
                    ~BalanceSubject.code.like(f"%{dot_pattern}.%"),
                )

        query = query.order_by(BalanceSubject.code)
        result = await session.execute(query)
        subjects = result.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "科目余额表"
    ws.append(["科目编码", "科目名称", "年初借方", "年初贷方",
                "本期借方", "本期贷方", "本年累计借方", "本年累计贷方",
                "期末借方", "期末贷方", "核算维度"])

    for s in subjects:
        ws.append(list(s))

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"科目余额表_{book_name}"
    if code_prefix:
        filename += f"_{code_prefix}"
    filename += ".xlsx"

    return _download_excel(stream, filename)