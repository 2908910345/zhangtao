import openpyxl
import math


def safe_num(v):
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        if math.isnan(v):
            return 0.0
        return float(v)
    try:
        return float(str(v).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0.0


def find_header_row(rows, keywords):
    for i, row in enumerate(rows):
        if i >= 20:
            break
        if not row or len(row) < 2:
            continue
        row_text = ''.join(str(c) for c in row)
        match_count = sum(1 for kw in keywords if kw in row_text)
        if match_count >= math.ceil(len(keywords) * 0.4):
            return i
    return -1


def _build_parent_category_map(header_row):
    filled = [''] * len(header_row)
    current = ''
    for c in range(len(header_row)):
        cell_str = str(header_row[c]).strip() if header_row[c] else ''
        if cell_str:
            current = cell_str
        filled[c] = current
    return filled


def parse_balance_sheet(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    header_row_idx = find_header_row(rows, ['科目编码', '科目名称', '余额'])
    if header_row_idx == -1:
        raise ValueError('无法识别科目余额表格式，请确保表头包含"科目编码"、"科目名称"等字段')

    header_row = [str(c) if c else '' for c in rows[header_row_idx]]
    data_start = header_row_idx + 1

    sub_header_row = []
    if header_row_idx + 1 < len(rows):
        next_row = rows[header_row_idx + 1]
        next_row_str = ''.join(str(c) for c in next_row if c)
        if '借方金额' in next_row_str and '贷方金额' in next_row_str:
            sub_header_row = [str(c) if c else '' for c in next_row]
            data_start = header_row_idx + 2

    code_col = None
    name_col = None
    dim_col = None
    year_start_debit_col = None
    year_start_credit_col = None
    period_debit_col = None
    period_credit_col = None
    year_total_debit_col = None
    year_total_credit_col = None
    end_debit_col = None
    end_credit_col = None

    for c, cell in enumerate(header_row):
        if '科目编码' in cell:
            code_col = c
        elif '科目名称' in cell:
            name_col = c
        elif '核算维度' in cell:
            dim_col = c

    if code_col is None or name_col is None:
        raise ValueError('未找到"科目编码"或"科目名称"列，请检查文件格式')

    if sub_header_row:
        filled_parent = _build_parent_category_map(header_row)

        parent_categories = {
            '年初余额': ('year_start_debit_col', 'year_start_credit_col'),
            '本期发生额': ('period_debit_col', 'period_credit_col'),
            '本年累计': ('year_total_debit_col', 'year_total_credit_col'),
            '期末余额': ('end_debit_col', 'end_credit_col'),
        }

        for c in range(len(sub_header_row)):
            sub = sub_header_row[c]
            parent = filled_parent[c] if c < len(filled_parent) else ''

            for cat, (debit_var, credit_var) in parent_categories.items():
                if cat in parent:
                    if '借方' in sub:
                        if debit_var == 'year_start_debit_col':
                            year_start_debit_col = c
                        elif debit_var == 'period_debit_col':
                            period_debit_col = c
                        elif debit_var == 'year_total_debit_col':
                            year_total_debit_col = c
                        elif debit_var == 'end_debit_col':
                            end_debit_col = c
                    elif '贷方' in sub:
                        if credit_var == 'year_start_credit_col':
                            year_start_credit_col = c
                        elif credit_var == 'period_credit_col':
                            period_credit_col = c
                        elif credit_var == 'year_total_credit_col':
                            year_total_credit_col = c
                        elif credit_var == 'end_credit_col':
                            end_credit_col = c
                    break
    else:
        for c, cell in enumerate(header_row):
            cl = cell
            if '年初' in cl:
                if '借方' in cl:
                    year_start_debit_col = c
                elif '贷方' in cl:
                    year_start_credit_col = c
            elif '本期' in cl:
                if '借方' in cl:
                    period_debit_col = c
                elif '贷方' in cl:
                    period_credit_col = c
            elif '本年' in cl:
                if '借方' in cl:
                    year_total_debit_col = c
                elif '贷方' in cl:
                    year_total_credit_col = c
            elif '期末' in cl:
                if '借方' in cl:
                    end_debit_col = c
                elif '贷方' in cl:
                    end_credit_col = c

    items = []
    for r in range(data_start, len(rows)):
        row = rows[r]
        if not row or len(row) < 2:
            continue
        code = str(row[code_col]).strip() if code_col < len(row) and row[code_col] is not None else ''
        name = str(row[name_col]).strip() if name_col < len(row) and row[name_col] is not None else ''
        if not code or not name:
            continue

        def get_col(col_idx):
            if col_idx is None or col_idx >= len(row):
                return 0
            return safe_num(row[col_idx])

        dim_val = ''
        if dim_col is not None:
            dim_val = str(row[dim_col]).strip() if dim_col < len(row) and row[dim_col] is not None else ''

        items.append({
            'code': code,
            'name': name,
            'year_start_debit': get_col(year_start_debit_col),
            'year_start_credit': get_col(year_start_credit_col),
            'period_debit': get_col(period_debit_col),
            'period_credit': get_col(period_credit_col),
            'year_total_debit': get_col(year_total_debit_col),
            'year_total_credit': get_col(year_total_credit_col),
            'end_debit': get_col(end_debit_col),
            'end_credit': get_col(end_credit_col),
            'dimension': dim_val,
        })

    wb.close()
    return items


def parse_journal_streaming(file_path):
    """流式解析序时账 Excel 文件，逐行 yield（使用 read_only 模式降低内存占用）"""
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    ws = wb.worksheets[0]

    header_keywords = ['凭证号', '科目编码', '摘要', '借方', '贷方']
    col_map = {}
    header_found = False
    row_idx = -1

    for row in ws.iter_rows(values_only=True):
        row_idx += 1

        if not header_found:
            if row_idx >= 20:
                wb.close()
                raise ValueError('无法识别序时账格式，请确保表头包含"凭证号"、"科目编码"、"借方"、"贷方"等字段')

            row_text = ''.join(str(c) for c in row)
            match_count = sum(1 for kw in header_keywords if kw in row_text)
            if match_count >= 3:
                header_row = [str(c) if c else '' for c in row]
                col_map = _build_journal_col_map(header_row)
                if 'voucher_no' not in col_map or 'code' not in col_map:
                    wb.close()
                    raise ValueError('未找到"凭证号"或"科目编码"列，请检查文件格式')
                header_found = True
            continue

        item = _parse_journal_row(row, col_map)
        if item:
            yield item

    wb.close()


def _build_journal_col_map(header_row):
    """构建序时账列名到列索引的映射"""
    col_map = {}
    for c, cell in enumerate(header_row):
        cl = cell
        if any(kw in cl for kw in ['核算组织', '公司', '单位']):
            col_map['org'] = c
        elif any(kw in cl for kw in ['记账日期', '业务日期', '日期']):
            col_map['date'] = c
        elif any(kw in cl for kw in ['期间', '会计期间']):
            col_map['period'] = c
        elif any(kw in cl for kw in ['凭证号', '凭证编号']):
            col_map['voucher_no'] = c
        elif any(kw in cl for kw in ['摘要', '业务摘要']):
            col_map['summary'] = c
        elif any(kw in cl for kw in ['科目编码', '科目代码', '科目编号']) and '科目名称' not in cl:
            col_map['code'] = c
        elif '科目名称' in cl:
            col_map['name'] = c
        elif '借方' in cl and '贷方' not in cl:
            col_map['debit'] = c
        elif '贷方' in cl:
            col_map['credit'] = c
        elif any(kw in cl for kw in ['核算维度', '辅助核算', '维度']):
            col_map['dimension'] = c
    return col_map


def _get_col_val(row, col_map, key):
    """从行数据中获取指定列的值"""
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return ''
    return str(row[idx]).strip() if row[idx] is not None else ''


def _parse_journal_row(row, col_map):
    """解析序时账单行数据"""
    voucher_no = _get_col_val(row, col_map, 'voucher_no')
    code = _get_col_val(row, col_map, 'code')
    if not voucher_no or not code:
        return None

    debit_val = 0.0
    credit_val = 0.0
    debit_idx = col_map.get('debit')
    credit_idx = col_map.get('credit')
    if debit_idx is not None and debit_idx < len(row):
        debit_val = safe_num(row[debit_idx])
    if credit_idx is not None and credit_idx < len(row):
        credit_val = safe_num(row[credit_idx])

    return {
        'org': _get_col_val(row, col_map, 'org'),
        'date': _get_col_val(row, col_map, 'date'),
        'period': _get_col_val(row, col_map, 'period'),
        'voucher_no': voucher_no,
        'summary': _get_col_val(row, col_map, 'summary'),
        'subject_code': code,
        'subject_name': _get_col_val(row, col_map, 'name'),
        'debit': debit_val,
        'credit': credit_val,
        'dimension': _get_col_val(row, col_map, 'dimension'),
    }